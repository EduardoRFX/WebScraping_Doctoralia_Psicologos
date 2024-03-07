from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side


def config_navegador():
    opcoes = webdriver.ChromeOptions()
    opcoes.add_experimental_option('detach', True)
    opcoes.add_experimental_option('excludeSwitches', ['enable-logging'])

    navegador = webdriver.Chrome(options=opcoes)
    navegador.get('https://www.doctoralia.com.br/pesquisa?q=Psicólogo&loc=São%20Paulo&filters%5Bspecializations%5D%5B%5D=76')
    
    return navegador


class Excel():
    def __init__(self):
        self.caminho = '<caminho do arquivo excel>'
        self.planilha = load_workbook(filename=self.caminho)
        self.sheetDados = self.planilha['<nome da sheet que receberá os dados>']

    def style_excel(self):  
        self.corTitulo = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
        self.corCelula = PatternFill(start_color='00808080', end_color='00808080', fill_type='solid')
        self.bordaStyle = Side(border_style='thin', color='000000')
        self.borda = Border(left=self.bordaStyle, top=self.bordaStyle, bottom=self.bordaStyle, right=self.bordaStyle)

    def config_excel(self):
        
        self.sheetDados['A1'] = 'Nomes'
        self.sheetDados['A1'].fill = self.corTitulo
        self.sheetDados['A1'].border = self.borda

        self.sheetDados['B1'] = 'Endereços'
        self.sheetDados['B1'].fill = self.corTitulo
        self.sheetDados['B1'].border = self.borda

        self.sheetDados['C1'] = 'Telefones'
        self.sheetDados['C1'].fill = self.corTitulo
        self.sheetDados['C1'].border = self.borda

        self.sheetDados.column_dimensions['A'].width = 37
        self.sheetDados.column_dimensions['B'].width = 111
        self.sheetDados.column_dimensions['C'].width = 35

    def transferir_excel(self, nome, endereco, telefone):

        ultimaLinha = len(self.sheetDados['A']) + 1

        self.sheetDados['A%s' % ultimaLinha] = nome
        self.sheetDados['A%s' % ultimaLinha].fill = self.corCelula
        self.sheetDados['A%s' % ultimaLinha].border = self.borda

        self.sheetDados['B%s' % ultimaLinha] = endereco
        self.sheetDados['B%s' % ultimaLinha].fill = self.corCelula
        self.sheetDados['B%s' % ultimaLinha].border = self.borda
        
        self.sheetDados['C%s' % ultimaLinha] = telefone
        self.sheetDados['C%s' % ultimaLinha].fill = self.corCelula
        self.sheetDados['C%s' % ultimaLinha].border = self.borda

        self.planilha.save(filename=self.caminho)